#!/usr/bin/env python3
"""
Extractor robusto de INVENTARIOS OXDA MAYO 2026..xlsx

- Hoja "Resumen Inventario": lee cada sección mensual y toma la columna de
  cobertura/último número por posición fija (última columna del encabezado).
- Hojas de almacén: busca la columna que contenga "INV MAYO" por posición y
  extrae el inventario de mayo para cada producto/código.
"""
import csv
import json
import re
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl no está instalado. Ejecuta: pip3 install openpyxl") from exc

XLSX_PATH = Path("INVENTARIOS OXDA MAYO 2026..xlsx")
OUTPUT_JSON = Path("inventario_xlsx_extraido.json")
OUTPUT_TS = Path("inventario_xlsx_extraido.ts")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    text = text.replace("$", "").replace("%", "").replace(",", "")
    text = re.sub(r"[\s\u00a0]+", "", text)
    text = text.replace("- ", "-")
    if text in ("", "-", "#div/0!", "#n/a", "#ref!", "#value!", "#name?", "#null!", "#num!", "#ref"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_column_by_header(ws, candidates: List[str], threshold: int = 4) -> Optional[int]:
    """Busca una columna cuyo encabezado (fila 1 o 2) coincida con algún candidato."""
    for row in (1, 2):
        for col in range(1, ws.max_column + 1):
            cell = normalize_text(ws.cell(row=row, column=col).value)
            for cand in candidates:
                if cand in cell or cell in cand:
                    return col
                # distancia Levenshtein simple para tolerar typos
                if levenshtein(cell, cand) <= threshold:
                    return col
    return None


def levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if len(a) < len(b):
        a, b = b, a
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(min(prev[j] + 1, curr[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = curr
    return prev[-1]


# ═══════════════════════════════════════════════════════════════════════════
# 1. RESUMEN INVENTARIO (secciones mensuales por cliente)
# ═══════════════════════════════════════════════════════════════════════════

def is_month_title_row(row: List[Any]) -> bool:
    for cell in row:
        if re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s*\.\s*\d{4}", normalize_text(cell)):
            return True
    return False


def extract_month_label(row: List[Any]) -> str:
    for cell in row:
        match = re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s*\.\s*(\d{4})", normalize_text(cell))
        if match:
            month = match.group(1).capitalize()
            year = match.group(2)
            return f"{month} {year}"
    return "Sin mes"


def is_header_row(row: List[Any]) -> bool:
    return any("cliente" in normalize_text(cell) and "bodega" in normalize_text(cell) for cell in row)


def is_total_row(row: List[Any]) -> bool:
    return any("importe total" in normalize_text(cell) for cell in row)


def is_client_row(row: List[Any]) -> bool:
    first = row[0] if row else None
    if first is None or str(first).strip() == "":
        return False
    text = normalize_text(first)
    if "cliente" in text or "importe total" in text:
        return False
    return True


def read_resumen_inventario(ws) -> List[Dict[str, Any]]:
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

    sections: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    header_indices: Dict[str, Optional[int]] = {}

    i = 0
    while i < len(rows):
        row = rows[i]
        if is_month_title_row(row):
            if current:
                sections.append(current)
            current = {"mes": extract_month_label(row), "clientes": [], "total": {}, "header_len": 0}
            header_indices = {}
            i += 1
            continue

        if current is None:
            i += 1
            continue

        if is_header_row(row):
            header_indices = {
                "cliente": 0,
                "valor_inventario": find_header_index(row, ["valo invent", "valo inventario", "valor inventario"]),
                "cajas_inventario": find_header_index(row, ["cajas invent", "cajas en inventario"]),
                "costo_promedio": find_header_index(row, ["costo prom", "costo promedio"]),
                "ventas_cajas_mes": find_header_index(row, ["vta cajas mes", "venta de cajas del mes"]),
                "llegadas": find_header_index(row, ["llegadas"]),
                "ventas_estimadas": find_header_index(row, ["ventas estimadas"]),
                "inv_mas_llegadas": find_header_index(row, ["inv actual + llegadas"]),
                "inv_menos_ventas": find_header_index(row, ["inv - ventas estimadas"]),
                "cobertura": find_header_index(row, ["cobertura"]),
            }
            if header_indices["cobertura"] is None and len(row) > 1:
                # última columna del encabezado por posición
                header_indices["cobertura"] = len(row) - 1
            current["header_len"] = len(row)
            i += 1
            continue

        if is_total_row(row):
            current["total"] = read_record(row, header_indices, current["header_len"])
            i += 1
            continue

        if is_client_row(row):
            rec = read_record(row, header_indices, current["header_len"])
            rec["cliente"] = str(row[0]).strip()
            current["clientes"].append(rec)
            i += 1
            continue

        i += 1

    if current:
        sections.append(current)
    return sections


def find_header_index(row: List[Any], candidates: List[str]) -> Optional[int]:
    """Busca el índice de columna cuyo encabezado contenga el candidato."""
    for idx, cell in enumerate(row):
        text = normalize_text(cell)
        if not text:
            continue
        for cand in candidates:
            if not cand:
                continue
            if cand in text or levenshtein(text, cand) <= 3:
                return idx
    return None


def read_record(row: List[Any], indices: Dict[str, Optional[int]], header_len: int) -> Dict[str, Any]:
    def get(key: str, default_idx: int):
        idx = indices.get(key)
        if idx is None and header_len > 0:
            idx = default_idx
        if idx is not None and idx < len(row):
            return parse_number(row[idx])
        return None

    return {
        "valor_inventario": get("valor_inventario", 1),
        "cajas_inventario": get("cajas_inventario", 2),
        "costo_promedio": get("costo_promedio", 3),
        "ventas_cajas_mes": get("ventas_cajas_mes", 4),
        "llegadas": get("llegadas", 6),
        "ventas_estimadas": get("ventas_estimadas", 7),
        "inv_mas_llegadas": get("inv_mas_llegadas", 8),
        "inv_menos_ventas": get("inv_menos_ventas", 9),
        "cobertura": get("cobertura", header_len - 1) if header_len > 0 else None,
        "ultimo_numero": get("cobertura", header_len - 1) if header_len > 0 else None,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 2. HOJAS DE ALMACÉN — columna INV MAYO por posición
# ═══════════════════════════════════════════════════════════════════════════

def read_almacen_sheet(ws, sheet_name: str) -> Dict[str, Any]:
    """
    Busca las columnas CÓDIGO, DESCRIPCIÓN e INV MAYO por posición en el
    encabezado y extrae el inventario de mayo de cada producto.
    """
    col_codigo: Optional[int] = None
    col_descripcion: Optional[int] = None
    col_inv_mayo: Optional[int] = None
    header_row: int = 1

    for r in range(1, min(6, ws.max_row + 1)):
        last_inv_mayo_in_row: Optional[int] = None
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c).value
            text = normalize_text(cell)
            if not text:
                continue
            if col_codigo is None and ("código" in text or "codigo" in text or text == "code"):
                col_codigo = c
                header_row = r
            if col_descripcion is None and ("descripción" in text or "descripcion" in text or "descrip" in text):
                col_descripcion = c
                header_row = r
            if "inv mayo" in text:
                # preferir la última aparición de INV MAYO en la fila (evita columnas duplicadas intermedias)
                last_inv_mayo_in_row = c
        if last_inv_mayo_in_row:
            col_inv_mayo = last_inv_mayo_in_row

    # Fallback: si no se encontró INV MAYO, buscar el último "INV <MES> 2026" de la hoja
    if col_inv_mayo is None:
        for r in range(1, min(6, ws.max_row + 1)):
            last_inv_2026: Optional[int] = None
            for c in range(1, ws.max_column + 1):
                text = normalize_text(ws.cell(row=r, column=c).value)
                if re.search(r"inv\s+\w+\s+2026", text):
                    last_inv_2026 = c
            if last_inv_2026:
                col_inv_mayo = last_inv_2026
                break

    productos = []
    if col_codigo is None:
        return {
            "hoja": sheet_name,
            "columna_codigo": col_codigo,
            "columna_descripcion": col_descripcion,
            "columna_inv_mayo": col_inv_mayo,
            "productos": productos,
        }

    for r in range(header_row + 1, ws.max_row + 1):
        codigo = ws.cell(row=r, column=col_codigo).value
        if codigo is None or str(codigo).strip() == "":
            continue
        descripcion = ws.cell(row=r, column=col_descripcion).value if col_descripcion else None
        inv_mayo = parse_number(ws.cell(row=r, column=col_inv_mayo).value) if col_inv_mayo else None
        productos.append({
            "codigo": str(codigo).strip(),
            "descripcion": str(descripcion).strip() if descripcion else "",
            "inv_mayo": inv_mayo,
            "columna_inv_mayo": col_inv_mayo,
        })

    return {
        "hoja": sheet_name,
        "columna_codigo": col_codigo,
        "columna_descripcion": col_descripcion,
        "columna_inv_mayo": col_inv_mayo,
        "productos": productos,
    }


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def determine_status(cobertura: Optional[float], ventas_estimadas: Optional[float]) -> str:
    if ventas_estimadas is not None and ventas_estimadas == 0:
        return "sin movimiento"
    if cobertura is None:
        return "sin movimiento"
    if cobertura < 3:
        return "alerta"
    if cobertura > 7:
        return "exceso"
    return "optimo"


def build_cliente_ts(c: Dict[str, Any]) -> str:
    cobertura = c.get("cobertura")
    ventas_est = c.get("ventas_estimadas")
    status = determine_status(cobertura, ventas_est)
    return (
        "    { "
        f"cliente: {json.dumps(c['cliente'], ensure_ascii=False)}, "
        f"valorInventario: {round(c.get('valor_inventario') or 0)}, "
        f"cajasInventario: {round(c.get('cajas_inventario') or 0)}, "
        f"costoPromedio: {round(c.get('costo_promedio') or 0)}, "
        f"ventasCajasMes: {round(c.get('ventas_cajas_mes') or 0)}, "
        f"llegadas: {round(c.get('llegadas') or 0)}, "
        f"ventasEstimadas: {round(c.get('ventas_estimadas') or 0)}, "
        f"invMasLlegadas: {round((c.get('inv_mas_llegadas') or 0) * 100) / 100}, "
        f"invMenosVentas: {round((c.get('inv_menos_ventas') or 0) * 100) / 100}, "
        f"cobertura: {cobertura if cobertura is not None else 0}, "
        f"status: {json.dumps(status)} "
        "},"
    )


def build_report_typescript(data: Dict[str, Any], meses_filtro: Optional[List[str]] = None) -> str:
    """
    Genera el módulo TS con los arreglos para el reporte.
    Si se pasan meses_filtro (en español, ej. ['enero', 'mayo']), solo incluye esos meses.
    """
    resumen = data.get("resumen_mensual", [])

    if meses_filtro:
        meses_norm = [m.lower()[:3] for m in meses_filtro]
        resumen = [s for s in resumen if s["mes"].split()[0].lower()[:3] in meses_norm]

    # Arreglo con todos los meses disponibles
    por_mes_lines = []
    for sec in resumen:
        clientes = "\n".join(build_cliente_ts(c) for c in sec["clientes"])
        por_mes_lines.append(
            f"  {{\n"
            f"    mes: {json.dumps(sec['mes'], ensure_ascii=False)},\n"
            f"    clientes: [\n{clientes}\n    ]\n"
            f"  }},"
        )

    # Alias de mayo para compatibilidad (el mes más reciente con datos)
    mayo = next((s for s in resumen if s["mes"].lower().startswith("mayo")), None)
    mayo_block = ""
    if mayo:
        clientes = "\n".join(build_cliente_ts(c) for c in mayo["clientes"])
        mayo_block = (
            "export const inventarioMayo2026 = [\n"
            + clientes
            + "\n];\n\n"
        )

    # Histórico mensual
    historico_lines = []
    meses_cortos = {
        "Diciembre": "Dic", "Enero": "Ene", "Febrero": "Feb", "Marzo": "Mar",
        "Abril": "Abr", "Mayo": "May", "Junio": "Jun", "Julio": "Jul",
        "Agosto": "Ago", "Septiembre": "Sep", "Octubre": "Oct", "Noviembre": "Nov",
    }
    for sec in resumen:
        total = sec.get("total", {})
        mes = sec["mes"].split()[0]
        corto = meses_cortos.get(mes, mes)
        anio = sec["mes"].split()[1]
        historico_lines.append(
            f"  {{ mes: \"{corto} {anio[-2:]}\", valor: {round(total.get('valor_inventario') or 0)}, cajas: {round(total.get('cajas_inventario') or 0)} }},"
        )

    return (
        mayo_block
        + "export const inventarioPorMes = [\n"
        + "\n".join(por_mes_lines)
        + "\n];\n\n"
        + "export const historicoMensual = [\n"
        + "\n".join(historico_lines)
        + "\n];\n"
    )


def parse_args() -> Any:
    import argparse
    parser = argparse.ArgumentParser(
        description="Extrae datos de INVENTARIOS OXDA MAYO 2026..xlsx"
    )
    parser.add_argument(
        "--mes",
        action="append",
        help="Mes a extraer (puede repetirse). Ej: --mes enero --mes mayo. Por defecto extrae todos.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    meses_filtro: Optional[List[str]] = args.mes

    if not XLSX_PATH.exists():
        print(f"No se encontró {XLSX_PATH}")
        return

    wb = load_workbook(XLSX_PATH, data_only=True)

    resumen = []
    if "Resumen Inventario" in wb.sheetnames:
        resumen = read_resumen_inventario(wb["Resumen Inventario"])

    almacenes = []
    for name in wb.sheetnames:
        if name in ("Resumen Inventario", "$$$$$", "Resumen Ventas", "Resumen T.C."):
            continue
        almacenes.append(read_almacen_sheet(wb[name], name))

    data = {
        "archivo": str(XLSX_PATH),
        "resumen_mensual": resumen,
        "almacenes": almacenes,
    }

    OUTPUT_JSON.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    OUTPUT_TS.write_text("export const inventarioXlsx = " + json.dumps(data, indent=2, ensure_ascii=False) + ";", encoding="utf-8")

    report_ts = Path("lib/inventario-data.ts")
    report_ts.write_text(build_report_typescript(data, meses_filtro=meses_filtro), encoding="utf-8")

    print(f"Resumen Inventario: {len(resumen)} secciones")
    for sec in resumen:
        print(f"\n{sec['mes']} ({len(sec['clientes'])} clientes)")
        for c in sec["clientes"]:
            print(f"  - {c['cliente']:<45} ultimo={c['ultimo_numero']} cobertura={c['cobertura']}")

    print(f"\nAlmacenes procesados: {len(almacenes)}")
    for alm in almacenes:
        col_label = alm['columna_inv_mayo'] if alm['columna_inv_mayo'] is not None else "N/A"
        print(f"  - {alm['hoja']:<25} col_inv_mayo={col_label:<3} productos={len(alm['productos'])}")
        if alm["productos"]:
            for p in alm["productos"][:3]:
                print(f"      {p['codigo']:<10} {p['descripcion'][:30]:<33} inv_mayo={p['inv_mayo']}")

    print(f"\nArchivos generados:")
    print(f"  - {OUTPUT_JSON}")
    print(f"  - {OUTPUT_TS}")
    print(f"  - {report_ts}")


if __name__ == "__main__":
    main()
