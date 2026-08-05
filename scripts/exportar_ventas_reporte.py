#!/usr/bin/env python3
"""Genera el CSV normalizado que consume la analítica desde el último Excel.

Uso:
    python3 scripts/exportar_ventas_reporte.py

El archivo generado conserva la estructura de la exportación original y se
versiona como corte auditable. No modifica el libro fuente.
"""

from pathlib import Path
import csv

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "2026 Reporte de Ventas Oxda al 24.06.2026.xlsx"
TARGET = ROOT / "data" / "ventas-reporte-2026.csv"


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"No se encontró el corte fuente: {SOURCE.name}")

    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["Reporte de Venta"]
    TARGET.parent.mkdir(exist_ok=True)

    with TARGET.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Tabla 1"])
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])

    print(f"Corte exportado: {TARGET.relative_to(ROOT)} ({sheet.max_row - 1} registros)")


if __name__ == "__main__":
    main()
